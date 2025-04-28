import pandas as pd
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

# Paths
json_folder = 'path_to_json_folder'
excel_folder = 'path_to_excel_files'
backup_folder = os.path.join(excel_folder, 'backup')
archive_folder = os.path.join(json_folder, 'archive')
sheet_name = 'utilisation'

# Columns to be updated
updatable_columns = [
    'analysis', 'design', 'test execution', 'regression', 'demo',
    'leave', 'downtime'
]

# Create backup and archive folders if they don’t exist
os.makedirs(backup_folder, exist_ok=True)
os.makedirs(archive_folder, exist_ok=True)

# Process each JSON file
for filename in os.listdir(json_folder):
    if filename.endswith('.json'):
        filepath = os.path.join(json_folder, filename)
        json_data = pd.read_json(filepath)
        json_data['date'] = pd.to_datetime(json_data['date'])

        for team_name, team_df in json_data.groupby('team'):
            excel_file = os.path.join(excel_folder, f"{team_name}.xlsx")
            if not os.path.exists(excel_file):
                print(f"Excel file for team '{team_name}' not found. Skipping.")
                continue

            # Backup the Excel file with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(
                backup_folder, f"{team_name}_{timestamp}.xlsx"
            )
            shutil.copy2(excel_file, backup_file)
            print(f"Backed up '{team_name}.xlsx' to '{backup_file}'")

            # Load workbook and worksheet
            wb = load_workbook(excel_file)
            if sheet_name not in wb.sheetnames:
                print(f"Sheet '{sheet_name}' not found in '{excel_file}'. Skipping.")
                continue

            ws = wb[sheet_name]

            # Load sheet to DataFrame for locating rows
            df_excel = pd.DataFrame(ws.values)
            df_excel.columns = df_excel.iloc[0]
            df_excel = df_excel[1:].copy()
            df_excel['date'] = pd.to_datetime(df_excel['date'])
            df_excel.reset_index(drop=True, inplace=True)

            # Map Excel headers to column letters
            header_row = 1
            column_map = {cell.value: cell.column_letter for cell in ws[header_row] if cell.value}

            # Update matching rows
            for _, row in team_df.iterrows():
                empid = row['empid']
                date = row['date']

                match = df_excel[(df_excel['empid'] == empid) & (df_excel['date'] == date)]
                if not match.empty:
                    excel_row = match.index[0] + 2
                    for col in updatable_columns:
                        if col in column_map and pd.notna(row[col]):
                            ws[f"{column_map[col]}{excel_row}"] = row[col]

            wb.save(excel_file)
            print(f"Updated Excel file: {excel_file}")

        # Move the JSON file to archive after processing
        archived_path = os.path.join(archive_folder, filename)
        shutil.move(filepath, archived_path)
        print(f"Archived JSON: {archived_path}")
